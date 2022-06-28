import os
import time
from pathlib import Path

import openpyxl

base_path = Path(__file__).parent


def readExcel(file_excel, name):
    wb = openpyxl.load_workbook(f'{base_path}/dataExcel_point/{file_excel}')
    sheet = wb[f'{name}']
    list_stt = []
    name_file = file_excel.rstrip('.xlsx').lstrip('DS 21-22_')
    n_name = name_file.replace(" ", "_")
    # loop to list danh sach STT from xlsx
    for i in range(1, 70):
        stt = sheet.cell(row=i, column=1)
        clasS = sheet.cell(row=i, column=5)
        if (stt.value) and (clasS.value):
            saint_name = sheet.cell(row=i, column=2)
            last_name = sheet.cell(row=i, column=3)
            first_name = sheet.cell(row=i, column=4)
            if stt.value == 'STT':
                continue
            dict = {}
            dict["SAINT_NAME"] = saint_name.value
            dict["LAST_NAME"] = last_name.value
            dict["FIRST_NAME"] = first_name.value
            dict["CLASS"] = clasS.value
            dict["OWNER"] = n_name
            list_stt.append(dict)
    return list_stt


def updateOrInsertExcel():
    name_files = []
    danhsach = []
    pat = f'{base_path}/dataExcel_point'
    objects = os.listdir(pat)
    files_file = [f for f in objects if os.path.isfile(os.path.join(pat, f))]
    for file in files_file:
        name_file = file.rstrip('.xlsx').lstrip('DS 21-22_')
        danhsach += readExcel(file, name_file)
        name_files.append(name_file)
    return name_files, danhsach


if __name__ == '__main__':
    t = time.time()
    updateOrInsertExcel()
    print("time out: ", time.time() - t)
